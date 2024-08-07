import argparse
import os
import yaml
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime


def process_yaml_files(directory):
    operators = defaultdict(set)
    skipped_rules = []
    rule_types = defaultdict(set)

    for filename in os.listdir(directory):
        if filename.endswith('.yaml') or filename.endswith('.yml'):
            try:
                with open(os.path.join(directory, filename), 'r', encoding='utf-8') as file:
                    try:
                        data = yaml.safe_load(file)
                    except yaml.YAMLError as e:
                        print(f"Error parsing {filename}: {e}")
                        skipped_rules.append((filename, str(e)))
                        continue

                    core_id = data.get('Core', {}).get('Id')
                    if core_id:
                        rule_type = data.get('Rule Type', 'Unknown')
                        rule_types[rule_type].add(core_id)

                        # Check the 'Check' section
                        check = data.get('Check', {})
                        if isinstance(check, dict):
                            for key, value in check.items():
                                if isinstance(value, list):
                                    for item in value:
                                        operator = item.get('operator')
                                        if operator:
                                            operators[operator].add(core_id)
                        elif isinstance(check, list):
                            for item in check:
                                operator = item.get('operator')
                                if operator:
                                    operators[operator].add(core_id)

                        # Check the 'Operations' section
                        operations = data.get('Operations', [])
                        for operation in operations:
                            operator = operation.get('operator')
                            if operator:
                                operators[operator].add(core_id)

            except UnicodeDecodeError:
                print(f"Encoding error in {filename}. Trying alternative encoding...")
                try:
                    with open(os.path.join(directory, filename), 'r', encoding='latin-1') as file:
                        content = file.read()
                    data = yaml.safe_load(content)
                    # Process data as before...
                    core_id = data.get('Core', {}).get('Id')
                    if core_id:
                        rule_type = data.get('Rule Type', 'Unknown')
                        rule_types[rule_type].add(core_id)

                        # Check the 'Check' section
                        check = data.get('Check', {})
                        if isinstance(check, dict):
                            for key, value in check.items():
                                if isinstance(value, list):
                                    for item in value:
                                        operator = item.get('operator')
                                        if operator:
                                            operators[operator].add(core_id)
                        elif isinstance(check, list):
                            for item in check:
                                operator = item.get('operator')
                                if operator:
                                    operators[operator].add(core_id)

                        # Check the 'Operations' section
                        operations = data.get('Operations', [])
                        for operation in operations:
                            operator = operation.get('operator')
                            if operator:
                                operators[operator].add(core_id)

                except Exception as e:
                    print(f"Failed to process {filename}: {e}")
                    skipped_rules.append((filename, str(e)))

    return operators, skipped_rules, rule_types


def write_results_to_excel(operators, skipped_rules, rule_types):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Operators"

    # Write headers
    sheet['A1'] = 'Operator'
    sheet['B1'] = 'Rule IDs'

    # Start writing from row 2
    row = 2
    for operator, rule_ids in operators.items():
        sheet.cell(row=row, column=1, value=operator)
        sheet.cell(row=row, column=2, value=', '.join(sorted(rule_ids)))
        row += 1

    # Set column widths and text wrapping for Operators sheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 100
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Create a new sheet for skipped rules
    skipped_sheet = wb.create_sheet(title="Skipped Rules")
    skipped_sheet['A1'] = 'Filename'
    skipped_sheet['B1'] = 'Error'

    # Write skipped rules
    for row, (filename, error) in enumerate(skipped_rules, start=2):
        skipped_sheet.cell(row=row, column=1, value=filename)
        skipped_sheet.cell(row=row, column=2, value=error)

    # Set column widths and text wrapping for Skipped Rules sheet
    skipped_sheet.column_dimensions['A'].width = 50
    skipped_sheet.column_dimensions['B'].width = 100
    for row in skipped_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Create a new sheet for rule types
    rule_type_sheet = wb.create_sheet(title="Rule Types")
    rule_type_sheet['A1'] = 'Rule Type'
    rule_type_sheet['B1'] = 'Core IDs'

    # Write rule types
    for row, (rule_type, core_ids) in enumerate(rule_types.items(), start=2):
        rule_type_sheet.cell(row=row, column=1, value=rule_type)
        rule_type_sheet.cell(row=row, column=2, value=', '.join(sorted(core_ids)))

    # Set column widths and text wrapping for Rule Types sheet
    rule_type_sheet.column_dimensions['A'].width = 30
    rule_type_sheet.column_dimensions['B'].width = 100
    for row in rule_type_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Generate filename with current date and time
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"rules_{current_time}.xlsx"

    wb.save(filename)
    return filename


def main():
    parser = argparse.ArgumentParser(description="Process YAML files and extract operators and rule IDs.")
    parser.add_argument("directory", help="Directory containing YAML files")
    args = parser.parse_args()

    operators, skipped_rules, rule_types = process_yaml_files(args.directory)
    output_file = write_results_to_excel(operators, skipped_rules, rule_types)
    print(f"Results written to {output_file}")
    print(f"Total rules processed: {sum(len(ids) for ids in rule_types.values())}")
    print(f"Total rules skipped: {len(skipped_rules)}")
    print(f"Total rule types: {len(rule_types)}")


if __name__ == "__main__":
    main()